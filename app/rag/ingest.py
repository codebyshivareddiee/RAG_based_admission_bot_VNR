"""
Document ingestion pipeline for Pinecone vector database.

Supports:
- Plain text files (.txt)
- PDF files (.pdf)
- Markdown files (.md)
- Excel files (.xlsx, .xls)
- CSV files (.csv)

Each chunk is stored with metadata:
  { "college": "VNRVJIET", "source": "...", "year": 2025 }

Usage:
    python -m app.rag.ingest --docs-dir ./docs
"""

from __future__ import annotations

import argparse
import hashlib
import os
import re
from pathlib import Path
from typing import Generator

from pinecone import Pinecone, ServerlessSpec
from openai import OpenAI

from app.config import get_settings

settings = get_settings()

# ── Clients ───────────────────────────────────────────────────
_openai_client: OpenAI | None = None
_pinecone_index = None


def _get_openai() -> OpenAI:
    global _openai_client
    if _openai_client is None:
        _openai_client = OpenAI(api_key=settings.OPENAI_API_KEY)
    return _openai_client


def _get_index():
    global _pinecone_index
    if _pinecone_index is None:
        pc = Pinecone(api_key=settings.PINECONE_API_KEY)

        # Create index if it doesn't exist
        existing = [idx.name for idx in pc.list_indexes()]
        if settings.PINECONE_INDEX_NAME not in existing:
            pc.create_index(
                name=settings.PINECONE_INDEX_NAME,
                dimension=1536,  # text-embedding-3-small
                metric="cosine",
                spec=ServerlessSpec(
                    cloud="aws",
                    region=settings.PINECONE_ENVIRONMENT,
                ),
            )
            print(f"✅  Created Pinecone index: {settings.PINECONE_INDEX_NAME}")

        _pinecone_index = pc.Index(settings.PINECONE_INDEX_NAME)
    return _pinecone_index


# ── Text extraction ───────────────────────────────────────────

def _read_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def _read_pdf_file(path: Path) -> str:
    """Extract text from PDF, with special handling for tables."""
    try:
        import pdfplumber
    except ImportError:
        raise ImportError("Install pdfplumber:  pip install pdfplumber")

    parts: list[str] = []
    with pdfplumber.open(str(path)) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            # Try extracting tables first
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    if not table:
                        continue
                    # First row as headers
                    headers = [str(c).strip() if c else "" for c in table[0]]
                    for row in table[1:]:
                        cells = [str(c).strip() if c else "" for c in row]
                        if not any(cells):
                            continue
                        entry_parts = []
                        for h, v in zip(headers, cells):
                            if v:
                                entry_parts.append(f"{h}: {v}")
                        if entry_parts:
                            parts.append(", ".join(entry_parts))

            # Also extract non-table text from the page
            text = page.extract_text() or ""
            # Remove text that's already captured in tables to avoid duplication
            if text.strip() and not tables:
                parts.append(text.strip())

    return "\n".join(parts)


def _read_excel_file(path: Path) -> str:
    """Extract text from Excel (.xlsx, .xls) – converts each sheet to readable text."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Install openpyxl:  pip install openpyxl")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        parts.append(f"## Sheet: {sheet_name}")

        # First row as headers
        headers = [str(c) if c is not None else "" for c in rows[0]]

        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            # Skip completely empty rows
            if not any(cells):
                continue
            # Format as "Header: Value" pairs for better RAG retrieval
            entry_parts = []
            for h, v in zip(headers, cells):
                if v:
                    entry_parts.append(f"{h}: {v}")
            if entry_parts:
                parts.append(", ".join(entry_parts))

    wb.close()
    return "\n".join(parts)


def _read_csv_file(path: Path) -> str:
    """Extract text from CSV files."""
    import csv

    parts: list[str] = []
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return ""

    headers = rows[0]
    for row in rows[1:]:
        if not any(row):
            continue
        entry_parts = []
        for h, v in zip(headers, row):
            if v:
                entry_parts.append(f"{h}: {v}")
        if entry_parts:
            parts.append(", ".join(entry_parts))

    return "\n".join(parts)


def _extract_text(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return _read_pdf_file(path)
    if ext in (".xlsx", ".xls"):
        return _read_excel_file(path)
    if ext == ".csv":
        return _read_csv_file(path)
    if ext in (".txt", ".md", ".markdown"):
        return _read_text_file(path)
    raise ValueError(f"Unsupported file type: {ext}")


# ── Chunking ──────────────────────────────────────────────────

def _section_aware_chunk(
    text: str,
    max_tokens: int = 350,
    overlap_tokens: int = 50,
) -> Generator[str, None, None]:
    """
    Section-aware chunking:
    1. Split by headings / double newlines first.
    2. Then split oversized sections by token count with overlap.
    """
    # Split on markdown headings or double newline
    sections = re.split(r"(?:\n\s*#{1,4}\s+|\n{2,})", text)
    sections = [s.strip() for s in sections if s.strip()]

    buffer: list[str] = []
    buffer_len = 0

    for section in sections:
        words = section.split()
        section_len = len(words)

        if buffer_len + section_len <= max_tokens:
            buffer.append(section)
            buffer_len += section_len
        else:
            # Flush buffer
            if buffer:
                yield " ".join(buffer)
            # If section itself exceeds max_tokens, split it
            if section_len > max_tokens:
                start = 0
                while start < section_len:
                    end = min(start + max_tokens, section_len)
                    yield " ".join(words[start:end])
                    start = end - overlap_tokens
            else:
                buffer = [section]
                buffer_len = section_len
                continue
            buffer = []
            buffer_len = 0

    if buffer:
        yield " ".join(buffer)


# ── Embedding ─────────────────────────────────────────────────

def _embed_texts(texts: list[str]) -> list[list[float]]:
    """Embed a batch of texts using OpenAI."""
    client = _get_openai()
    response = client.embeddings.create(
        input=texts,
        model=settings.OPENAI_EMBEDDING_MODEL,
    )
    return [item.embedding for item in response.data]


# ── Ingestion ─────────────────────────────────────────────────

def ingest_file(
    path: Path,
    source_label: str = "document",
    year: int = 2025,
    batch_size: int = 50,
) -> int:
    """
    Ingest a single document into Pinecone.

    Returns the number of chunks upserted.
    """
    text = _extract_text(path)
    chunks = list(_section_aware_chunk(text))

    if not chunks:
        print(f"⚠️  No content extracted from {path.name}")
        return 0

    index = _get_index()
    total = 0

    for i in range(0, len(chunks), batch_size):
        batch = chunks[i : i + batch_size]
        embeddings = _embed_texts(batch)

        vectors = []
        for j, (chunk, emb) in enumerate(zip(batch, embeddings)):
            chunk_id = hashlib.sha256(
                f"{path.name}:{i + j}:{chunk[:64]}".encode()
            ).hexdigest()[:32]

            vectors.append(
                {
                    "id": chunk_id,
                    "values": emb,
                    "metadata": {
                        "college": settings.COLLEGE_SHORT_NAME,
                        "source": source_label,
                        "year": year,
                        "filename": path.name,
                        "chunk_index": i + j,
                        "text": chunk[:2000],  # Pinecone metadata limit
                    },
                }
            )

        index.upsert(vectors=vectors)
        total += len(vectors)

    print(f"✅  Ingested {total} chunks from {path.name}")
    return total


def ingest_directory(
    docs_dir: str | Path,
    source_label: str = "document",
    year: int = 2025,
) -> int:
    """Ingest all supported files from a directory."""
    docs_path = Path(docs_dir)
    if not docs_path.is_dir():
        raise FileNotFoundError(f"Directory not found: {docs_dir}")

    supported = {".txt", ".md", ".markdown", ".pdf", ".xlsx", ".xls", ".csv"}
    total = 0

    for fpath in sorted(docs_path.iterdir()):
        if fpath.suffix.lower() in supported:
            total += ingest_file(fpath, source_label, year)

    print(f"\n🎉  Total chunks ingested: {total}")
    return total


# ── CLI ───────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Ingest documents into Pinecone")
    parser.add_argument(
        "--docs-dir",
        type=str,
        default="docs",
        help="Path to directory containing documents to ingest",
    )
    parser.add_argument("--source", type=str, default="document")
    parser.add_argument("--year", type=int, default=2025)
    args = parser.parse_args()

    ingest_directory(args.docs_dir, args.source, args.year)


if __name__ == "__main__":
    main()
